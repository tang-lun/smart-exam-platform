import csv
import io
import re

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.models.question import Question, QuestionSource
from app.models.user import User
from app.schemas.question import (
    QuestionGenerateRequest,
    QuestionListResponse,
    QuestionManualCreate,
    QuestionResponse,
    QuestionUpdateRequest,
)
from app.services.ai_service import generate_analysis, generate_questions, validate_question
from app.services.auth_service import get_current_user

router = APIRouter(prefix="/api/questions", tags=["questions"])

# 清理 AI 可能在 LaTeX 公式内生成的不规范填空标记
_TEXT_BLANK_RE = re.compile(r'\\text\{_{3,}\}')


def _clean_stem(text: str) -> str:
    """将 LaTeX 公式内的 \\text{______} 替换为普通文本下划线。"""
    return _TEXT_BLANK_RE.sub('______', text)

VALID_TYPES = {"choice": "选择题", "fill_blank": "填空题", "calculation": "计算题", "proof": "证明题"}
VALID_DIFFICULTIES = {"easy": "基础", "medium": "中等", "hard": "较难"}
VALID_GRADES = {"grade_7": "初一", "grade_8": "初二", "grade_9": "初三"}

# 中文→英文 反向映射
TYPE_CN2EN = {v: k for k, v in VALID_TYPES.items()}
DIFF_CN2EN = {v: k for k, v in VALID_DIFFICULTIES.items()}
GRADE_CN2EN = {v: k for k, v in VALID_GRADES.items()}


def _normalize_type(val: str) -> str:
    val = val.strip()
    if val in VALID_TYPES:
        return val
    if val in TYPE_CN2EN:
        return TYPE_CN2EN[val]
    # 模糊匹配
    for cn, en in TYPE_CN2EN.items():
        if cn in val or val in cn:
            return en
    return val.lower()


def _normalize_diff(val: str) -> str:
    val = val.strip()
    if val in VALID_DIFFICULTIES:
        return val
    if val in DIFF_CN2EN:
        return DIFF_CN2EN[val]
    for cn, en in DIFF_CN2EN.items():
        if cn in val or val in cn:
            return en
    return val.lower()


def _normalize_grade(val: str) -> str:
    val = val.strip()
    if val in VALID_GRADES:
        return val
    if val in GRADE_CN2EN:
        return GRADE_CN2EN[val]
    for cn, en in GRADE_CN2EN.items():
        if cn in val or val in cn:
            return en
    return val.lower()
COLUMN_ALIASES = {
    "题型": "type", "题目类型": "type", "类型": "type",
    "题干": "stem", "题目": "stem", "题目内容": "stem",
    "选项a": "option_a", "选项A": "option_a", "a": "option_a", "A": "option_a",
    "选项b": "option_b", "选项B": "option_b", "b": "option_b", "B": "option_b",
    "选项c": "option_c", "选项C": "option_c", "c": "option_c", "C": "option_c",
    "选项d": "option_d", "选项D": "option_d", "d": "option_d", "D": "option_d",
    "答案": "answer", "正确答案": "answer",
    "解析": "answer_analysis", "答案解析": "answer_analysis", "解题过程": "answer_analysis",
    "知识点": "knowledge_points", "考查知识点": "knowledge_points",
    "难度": "difficulty", "题目难度": "difficulty",
    "年级": "grade_level", "学段": "grade_level", "适用年级": "grade_level",
}


def _visible(user: User | None):
    """返回当前用户可见的题目过滤条件（严格按用户隔离）。"""
    if user is None:
        return Question.owner_id == None  # noqa: E711
    return Question.owner_id == user.id


@router.post("/generate", response_model=list[QuestionResponse], status_code=201)
def generate(
    req: QuestionGenerateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """AI 生成题目并存入题库。"""
    raw_questions = generate_questions(
        knowledge_points=req.knowledge_points,
        question_type=req.question_type,
        count=req.count,
        difficulty=req.difficulty,
        grade_level=req.grade_level,
    )

    created = []
    for q in raw_questions:
        stem = _clean_stem(q["stem"].strip())
        existing = db.query(Question).filter(
            Question.stem == stem, _visible(current_user)
        ).first()
        if existing:
            continue
        question = Question(
            type=q.get("type", req.question_type),
            subject="math",
            grade_level=req.grade_level,
            knowledge_points=q.get("knowledge_points", req.knowledge_points),
            difficulty=q.get("difficulty", req.difficulty),
            stem=stem,
            options=q.get("options"),
            answer=q["answer"],
            answer_analysis=q.get("answer_analysis"),
            source=QuestionSource.ai_generated,
            owner_id=current_user.id,
        )
        db.add(question)
        created.append(question)

    db.commit()
    for q in created:
        db.refresh(q)

    return [QuestionResponse.from_question(q, current_user.id) for q in created]


@router.post("/manual", response_model=QuestionResponse, status_code=201)
def create_manual(
    req: QuestionManualCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """手动创建题目，AI 审核把关。"""
    q_data = req.model_dump()

    stem = _clean_stem(q_data["stem"].strip())
    if db.query(Question).filter(Question.stem == stem, _visible(current_user)).first():
        raise HTTPException(status_code=400, detail="题库中已存在相同题干的题目")

    result = validate_question(q_data)
    if not result["valid"]:
        raise HTTPException(status_code=400, detail=f"AI 审核不通过：{result['feedback']}")

    analysis = (q_data.get("answer_analysis") or "").strip()
    if not analysis:
        analysis = generate_analysis(q_data)

    question = Question(
        type=q_data["type"],
        subject="math",
        grade_level=q_data.get("grade_level", "grade_7"),
        knowledge_points=q_data["knowledge_points"],
        difficulty=q_data.get("difficulty", "medium"),
        stem=stem,
        options=q_data.get("options"),
        answer=q_data["answer"],
        answer_analysis=analysis or None,
        source=QuestionSource.manual,
        owner_id=current_user.id,
    )
    db.add(question)
    db.commit()
    db.refresh(question)
    return QuestionResponse.from_question(question, current_user.id)


@router.get("", response_model=QuestionListResponse)
def list_questions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: str = Query("", description="搜索题干关键词"),
    question_type: str = Query("", description="按题型筛选"),
    difficulty: str = Query("", description="按难度筛选"),
    grade_level: str = Query("", description="按学段筛选"),
    knowledge_point: str = Query("", description="按知识点筛选"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """题目列表，支持搜索和筛选。"""
    query = db.query(Question).filter(_visible(current_user))

    if keyword:
        query = query.filter(Question.stem.contains(keyword))
    if question_type:
        query = query.filter(Question.type == question_type)
    if difficulty:
        query = query.filter(Question.difficulty == difficulty)
    if grade_level:
        query = query.filter(Question.grade_level == grade_level)
    if knowledge_point:
        # SQLite JSON 列存储为文本，用 LIKE 匹配 JSON 数组中的字符串
        query = query.filter(Question.knowledge_points.like(f'%"{knowledge_point}"%'))

    total = query.count()
    items = query.order_by(Question.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return QuestionListResponse(
        items=[QuestionResponse.from_question(q, current_user.id) for q in items],
        total=total, page=page, page_size=page_size,
    )


@router.post("/validate")
def validate_manual_question(
    req: QuestionManualCreate,
    current_user: User = Depends(get_current_user),
):
    """AI 审核题目，不保存。"""
    result = validate_question(req.model_dump())
    return result


@router.get("/favorites/list", response_model=QuestionListResponse)
def list_favorites(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """收藏题目列表。"""
    from sqlalchemy import text

    query = db.query(Question).filter(
        text("EXISTS (SELECT 1 FROM json_each(questions.favorited_by) WHERE value = :uid)").bindparams(uid=current_user.id),
        _visible(current_user),
    )
    total = query.count()
    items = query.order_by(Question.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return QuestionListResponse(
        items=[QuestionResponse.from_question(q, current_user.id) for q in items],
        total=total, page=page, page_size=page_size,
    )


@router.get("/{question_id}", response_model=QuestionResponse)
def get_question(
    question_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取单道题目详情。"""
    question = db.query(Question).filter(
        Question.id == question_id, _visible(current_user)
    ).first()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")
    return QuestionResponse.from_question(question, current_user.id)


@router.put("/{question_id}", response_model=QuestionResponse)
def update_question(
    question_id: int,
    req: QuestionUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """编辑题目。"""
    question = db.query(Question).filter(
        Question.id == question_id, _visible(current_user)
    ).first()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")

    update_data = req.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(question, key, value)

    db.commit()
    db.refresh(question)
    return QuestionResponse.from_question(question, current_user.id)


@router.post("/{question_id}/favorite")
def toggle_favorite(
    question_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """切换收藏状态（按用户隔离）。"""
    question = db.query(Question).filter(
        Question.id == question_id, _visible(current_user)
    ).first()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")
    fav_list = list(question.favorited_by or [])
    uid = current_user.id
    if uid in fav_list:
        fav_list.remove(uid)
    else:
        fav_list.append(uid)
    question.favorited_by = fav_list
    db.commit()
    return {"is_favorited": uid in fav_list}


@router.post("/import")
def import_questions(
    file: UploadFile,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """批量导入题目（支持 CSV / Excel .xlsx）。"""
    filename = (file.filename or "").lower()
    if not filename.endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .csv / .xlsx / .xls 格式")

    # 解析文件内容
    content = file.file.read()
    rows = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip() if h else "" for h in row]
                continue
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = str(val).strip() if val is not None else ""
            if any(v for v in row_dict.values()):
                rows.append(row_dict)
        wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据行")

    # 列名映射
    def map_column(col_name: str) -> str:
        return COLUMN_ALIASES.get(col_name.strip(), col_name.strip().lower().replace(" ", "_"))

    success = 0
    errors = []

    for r_idx, raw_row in enumerate(rows):
        row = {map_column(k): v for k, v in raw_row.items() if k}

        stem = _clean_stem((row.get("stem") or "").strip())
        if not stem:
            errors.append({"row": r_idx + 2, "error": "题干为空"})
            continue

        qtype = _normalize_type(row.get("type", "choice"))
        if qtype not in VALID_TYPES:
            errors.append({"row": r_idx + 2, "error": f"题型无效：{row.get('type', '')}，应为 选择题/填空题/计算题/证明题 或 choice/fill_blank/calculation/proof"})
            continue

        difficulty = _normalize_diff(row.get("difficulty", "medium"))
        if difficulty not in VALID_DIFFICULTIES:
            difficulty = "medium"

        grade = _normalize_grade(row.get("grade_level", "grade_7"))
        if grade not in VALID_GRADES:
            grade = "grade_7"

        # 知识点：逗号/中文逗号分隔
        kp_raw = row.get("knowledge_points", "").strip()
        knowledge_points = [k.strip() for k in kp_raw.replace("，", ",").split(",") if k.strip()] if kp_raw else []

        # 选项
        options = None
        if qtype == "choice":
            opts = []
            for key in ("option_a", "option_b", "option_c", "option_d"):
                val = row.get(key, "").strip()
                if val:
                    label = key[-1].upper()
                    opts.append(f"{label}. {val}")
            if len(opts) < 2:
                errors.append({"row": r_idx + 2, "error": "选择题至少需要2个选项"})
                continue
            options = opts

        answer = row.get("answer", "").strip()
        if not answer:
            errors.append({"row": r_idx + 2, "error": "答案为空"})
            continue

        analysis = row.get("answer_analysis", "").strip() or None

        # 去重（仅限当前用户题库内）
        if db.query(Question).filter(Question.stem == stem, _visible(current_user)).first():
            errors.append({"row": r_idx + 2, "error": f"题库已存在：{stem[:30]}..."})
            continue

        try:
            question = Question(
                type=qtype,
                subject="math",
                grade_level=grade,
                knowledge_points=knowledge_points,
                difficulty=difficulty,
                stem=stem,
                options=options,
                answer=answer,
                answer_analysis=analysis,
                source=QuestionSource.imported,
                owner_id=current_user.id,
            )
            db.add(question)
            success += 1
        except Exception as e:
            errors.append({"row": r_idx + 2, "error": str(e)})

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"数据保存失败：{str(e)}")

    return {
        "success": success,
        "errors": errors,
        "total": len(rows),
    }


@router.get("/template/download")
def download_import_template():
    """下载导入模板（CSV）。"""
    from fastapi.responses import StreamingResponse

    header = "题型,题干,选项A,选项B,选项C,选项D,答案,解析,知识点,难度,年级\n"
    example = (
        "选择题,计算 $-5 + 3$ 的结果,"
        "A. $-2$,B. $2$,C. $-8$,D. $8$,"
        "A,有理数加法法则：异号两数相加取绝对值较大数的符号，并用较大绝对值减去较小绝对值,"
        "有理数的运算,基础,初一\n"
    )
    content = header + example
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''题目导入模板.csv"},
    )


@router.delete("/{question_id}", status_code=204)
def delete_question(
    question_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """删除题目，并清理试卷中的引用。"""
    from app.models.exam import Exam

    question = db.query(Question).filter(
        Question.id == question_id, _visible(current_user)
    ).first()
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")

    db.delete(question)
    db.flush()

    # 从当前用户的试卷中移除该题目引用（不重排 ID）
    exams = db.query(Exam).filter(
        Exam.question_ids.isnot(None),
        Exam.owner_id == current_user.id,
    ).all()
    for exam in exams:
        qids = exam.question_ids or []
        if question_id in qids:
            deleted_pos = qids.index(question_id)
            exam.question_ids = [qid for qid in qids if qid != question_id]
            # 移除对应位置的分值，后续位置前移
            old_scores = exam.question_scores or {}
            new_scores = {}
            for pos_str, score in old_scores.items():
                pos = int(pos_str)
                if pos < deleted_pos:
                    new_scores[pos_str] = score
                elif pos > deleted_pos:
                    new_scores[str(pos - 1)] = score
            exam.question_scores = new_scores

    db.commit()
